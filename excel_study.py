import os
import openpyxl
import random

#파일을 읽어들일 위치 지정
#cwd = os.getcwd() 현재 프로그램이 있는 위치를 반환한다.
cwd = "/Users/jangjumyeong/Desktop/학습공동체(토익)"
filename = "a.xlsx"
filepath = os.path.join(cwd,filename)

#워크북 객체
wb = openpyxl.load_workbook(filepath)
print(wb)
print(type(wb))
print(wb.sheetnames)

#시트 객체
ws = wb['Sheet1']

#열행 객체를 불러온다.
cols = tuple(ws.columns)
rows = tuple(ws.rows)

#단어 저장 리스트이다.
word = list()

maxNum = 0
for col_idx,col in enumerate(rows):
    print(col[0].value,col[1].value)
    word.append( (col[0].value,col[1].value) )
    maxnum = col_idx

print("리스트 1번")
print(word[0])

random.shuffle(word)


print("섞은 리스트 1번")
print(word[0])

#내부의 값을 변경할 수 있다.
a3 = cols[0][2]
print(a3.value)
a3.value = 'row1'
print(a3.value)

#랜덤 문제를 엑셀에 할당한다.
for col_idx,col in enumerate(rows):
    intoSheet1 = rows[col_idx][0]
    intoSheet2 = rows[col_idx][1]
    if maxnum/2 > col_idx:
        intoSheet1.value = word[col_idx][0]
        #  정답 지우기
        intoSheet2.value = ''
    else:
        intoSheet1.value = word[col_idx][1]
        # #  정답 지우기
        intoSheet2.value = ''

#답지 생성하기
ws = wb.create_sheet("정답")

for col in range(maxnum+1):
    if maxnum/2 > col:
        ws.cell(row= col+1 ,column= 1, value= word[col][1])
        #  정답 지우기
        #ws.cell(row= col+1 ,column= 2, value= word[col][0])
    else:
        ws.cell(row= col+1 ,column= 1, value= word[col][0])
        #  정답 지우기
        #ws.cell(row= col+1 ,column= 2, value= word[col][1])




#워크북의 변경 사항을 저장 한다.
wb.save(os.path.join(cwd,"out.xlsx"))

#word 는 단어를 모아둔 것이다.
word = list()

active_sheet = wb.active
print(active_sheet)
